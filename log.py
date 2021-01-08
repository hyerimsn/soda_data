import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)) + '/app')))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django
import pandas as pd
from openpyxl import load_workbook
django.setup()
from search.models import Log

log_wb = load_workbook(r"C:\Users\USER\Desktop\pusan_parliament\searchengine\291~292 의회.xlsx", data_only=True)
log_ws = log_wb['291~292']

result = []
for i in range(log_ws.max_row):
    item_obj = {
        'person': log_ws['A'][i].value,
        'status': log_ws['B'][i].value,
        'ment': log_ws['C'][i].value,
        'date': log_ws['D'][i].value
    }
    result.append(item_obj)
    
for i in result:
    Log(person = i['person'],
        status = i['status'],
        ment = i['ment'],
        date = i['date']).save()