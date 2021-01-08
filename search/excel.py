import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)) + '/app')))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()

import django
from search.models import News
from openpyxl import load_workbook
django.setup()
news_wb = load_workbook(r"C:\Users\USER\Desktop\pusan_parliament\searchengine\NewsResult_20191225-20201225.xlsx", data_only=True)
news_ws = news_wb['sheet']

result = []

for i in range(news_ws.max_row):
    item_obj = {
        'title': news_ws['E'][i].value,
        'reporter': news_ws['D'][i].value,
        'media': news_ws['C'][i].value,
        'body': news_ws['Q'][i].value,
        'link': news_ws['R'][i].value,
        'date' : news_ws['B'][i].value,
    }

    result.append(item_obj)

for i in result:
    News(title = i['title'],
        reporter = i['reporter'],
        media = i['media'],
        body = i['body'],
        link = i['link'],
        date = i['date']).save()
